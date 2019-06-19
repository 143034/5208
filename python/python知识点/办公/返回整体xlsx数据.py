'''
返回整体xlsx数据
'''
from openpyxl.reader.excel import load_workbook
def readxlsFile(path):
    dic = {}
    file = load_workbook(filename=path)
    #print(file.get_sheet_names())
    sheets = file.get_sheet_names()
    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        #一张表的所有数据
        sheetInfo = []
        for lineNum in range(1, sheet.max_row +1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum, column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)

        dic[sheetName] = sheetInfo
    return dic
path = r'C:\Users\Administrator\Desktop\文件工作区\workspace4\20190530智慧树平台《形势与政策》未报到及学习进度0学生名单.xlsx'
dic = readxlsFile(path)
print(dic)