'''
读取xlsx文件
'''
from openpyxl.reader.excel import load_workbook
def readxlsFile(path):
    file = load_workbook(filename=path)
    #print(file.get_sheet_names())
    sheets = file.get_sheet_names()
    sheet = file.get_sheet_by_name(sheets[0])
    print(sheet.max_row)
    #print(sheet.max_column)
    #print(sheet.title)
    for lineNum in range(1,sheet.max_row +1):
        #print(lineNum)
        lineList = []
        for columnNum in range(1, sheet.max_column +1):
            value = sheet.cell(row=lineNum,column=columnNum).value
            if value != None:

                lineList.append(value)
        print(lineList)
path = r'C:\Users\Administrator\Desktop\文件工作区\workspace4\20190530智慧树平台《形势与政策》未报到及学习进度0学生名单.xlsx'
readxlsFile(path)